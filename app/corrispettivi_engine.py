import csv
import os
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

from app.engine import get_date_chunks
from app.security import get_ca_bundle

TIPI_ELENCO = ["RT", "DA", "MC", "CA"]


def date_to_rest(date_str: str) -> str:
    d = datetime.strptime(date_str, "%d/%m/%Y")
    return d.strftime("%d%m%Y")


def parse_amount(val: Any) -> Optional[float]:
    """
    Converte un valore importo in float.
    Supporta stringhe formato italiano (1.234,56) e inglese (1234.56)
    e numeri già float/int (es. da API JSON).

    Restituisce sempre un float pulito (mai stringa) o None se non valido.
    """
    if val is None:
        return None
    # Se è già un numero, restituiscilo direttamente
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return None
    try:
        cleaned = val.replace("+", "").strip()
        if "," in cleaned:
            # Formato italiano: 1.234,56 → rimuovi punti, cambia virgola in punto
            cleaned = cleaned.replace(".", "").replace(",", ".")
        return float(cleaned)
    except (ValueError, AttributeError):
        return None


def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)


def _extract_imponibile(record: Dict[str, Any]) -> Optional[float]:
    if record.get("importoParzialeTotale") is not None:
        return parse_amount(record["importoParzialeTotale"])
    return parse_amount(record.get("ammontareTotale"))


def _aliquote_map(riepilogo: List[Dict[str, Any]]) -> Dict[str, Optional[float]]:
    """
    Estrae l'imponibile + imposta per ogni aliquota IVA dal riepilogo.

    Restituisce un dict {aliquota: totale} dove totale = imponibile + imposta
    (importo comprensivo di IVA), incasellato nella colonna corretta
    in base all'aliquota applicata.
    """
    ali: Dict[str, Optional[float]] = {"4": None, "5": None, "10": None, "22": None, "ESENTE": None}
    for r in riepilogo:
        ali_val = r.get("aliquotaIva")
        if ali_val is not None:
            key = str(int(ali_val)) if ali_val == int(ali_val) else str(ali_val)
            if key in ali:
                imponibile = parse_amount(r.get("imponibile")) or 0.0
                imposta = parse_amount(r.get("imposta")) or 0.0
                ali[key] = imponibile + imposta
        natura = r.get("natura", "")
        if natura and ali_val is None:
            imponibile = parse_amount(r.get("imponibile")) or 0.0
            imposta = parse_amount(r.get("imposta")) or 0.0
            ali["ESENTE"] = imponibile + imposta
    return ali


_STANDARD_ALIQUOTE = {4, 5, 10, 22}


def _compute_aliquota(imponibile: float, imposta: float) -> Optional[int]:
    """Calcola l'aliquota IVA dal rapporto imposta/imponibile.

    Tolleranza adattiva: per importi piccoli (< 10€) la tolleranza
    viene aumentata per compensare arrotondamenti al centesimo.
    Per importi >= 10€, tolleranza stretta (0.1%).
    """
    if not imponibile or imponibile == 0:
        return None
    ratio = imposta / imponibile * 100
    # Tolleranza adattiva per compensare arrotondamenti al centesimo
    if imponibile < 1.0:
        toll = 10.0    # Molto permissivo sotto 1€
    elif imponibile < 10.0:
        toll = 2.0     # Permissivo tra 1€ e 10€
    else:
        toll = 0.1     # Stretto sopra 10€
    for std in sorted(_STANDARD_ALIQUOTE, reverse=True):
        if abs(ratio - std) < toll:
            return std
    return None


def _scomponi_atipico(
    imponibile: float,
    imposta: float,
) -> Dict[str, Optional[float]]:
    """Tenta di scomporre un record con aliquota mista.

    Euristico: prova ogni aliquota standard (22 → 10 → 5 → 4).
    Per ogni aliquota calcola:
      imponibile_aliquota = imposta / aliquota
      esente = imponibile_totale - imponibile_aliquota
    Se esente >= 0 ed è "tondo" (multiplo di 0.05 o differenza < 1%)
    la scomposizione è considerata valida.

    Returns:
        dict con chiavi "4", "5", "10", "22", "ESENTE"
        contenente il totale (imponibile+imposta) per ogni aliquota,
        oppure None per le aliquote non presenti.
    """
    ali: Dict[str, Optional[float]] = {
        "4": None, "5": None, "10": None, "22": None, "ESENTE": None,
    }

    for ali_val in (22, 10, 5, 4):
        if imposta == 0:
            # Imposta zero → tutto esente
            ali["ESENTE"] = imponibile
            break
        imponibile_aliquota = imposta / (ali_val / 100.0)
        if imponibile_aliquota > imponibile:
            continue  # Aliquota troppo bassa (imponibile > totale)

        esente = imponibile - imponibile_aliquota
        # Verifica se esente è "tondo" o se è una parte ragionevole
        esente_arrotondato = round(esente, 2)
        diff = abs(esente - esente_arrotondato)

        # Criterio: esente tondo (diff < 0.01) OPPURE
        # esente è una parte ragionevole del totale (> 1%)
        if diff < 0.01 or esente / max(imponibile, 0.01) > 0.01:
            totale_aliquota = imponibile_aliquota + imposta
            totale_aliquota = round(totale_aliquota, 2)
            esente_round = round(esente, 2)

            # Solo se esente è >= 0
            if esente_round >= 0:
                ali[str(ali_val)] = totale_aliquota
                if esente_round > 0:
                    ali["ESENTE"] = esente_round
                return ali

    # Fallback: se nessuna scomposizione funziona,
    # metti TUTTO nell'aliquota più vicina
    ratio = imposta / max(imponibile, 0.01) * 100 if imposta else 0
    for std in sorted(_STANDARD_ALIQUOTE, reverse=True):
        if abs(ratio - std) < 5.0:
            ali[str(std)] = round(imponibile + imposta, 2)
            return ali

    # Ultimo fallback: tutto a 22%
    ali["22"] = round(imponibile + imposta, 2)
    return ali


def _extract_dettaglio_iva(record: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Estrae il dettaglio IVA da un record corrispettivi.

    Strategie:
      1. Array 'riepilogo'/'dettaglioIva'/'aliquote'.
      2. Campo 'aliquotaIva'/'aliquota' a livello top.
      3. Calcolo inverso: rate = round(imposta / imponibile).
    """
    for key in ("riepilogo", "dettaglioIva", "aliquote", "linee"):
        val = record.get(key)
        if isinstance(val, list) and val:
            return val
    for ali_key in ("aliquotaIva", "aliquota"):
        ali_val = record.get(ali_key)
        if ali_val is not None:
            imponibile = (record.get("imponibile") or record.get("importoParzialeTotale") or record.get("ammontareTotale"))
            imposta = record.get("imposta") or record.get("impostaTotale")
            return [{"aliquotaIva": ali_val, "imponibile": imponibile, "imposta": imposta}]
    # 3. Calcolo rate da imposta/imponibile (API senza aliquota esplicita)
    imponibile = record.get("importoParzialeTotale") or record.get("ammontareTotale")
    imposta = record.get("impostaTotale")
    if imponibile and imposta:
        imponibile_f = parse_amount(imponibile)
        imposta_f = parse_amount(imposta)
        if imponibile_f and imponibile_f > 0:
            ali = _compute_aliquota(imponibile_f, imposta_f)
            if ali is not None:
                return [{"aliquotaIva": ali, "imponibile": str(imponibile_f), "imposta": str(imposta_f)}]
    return []


def _extract_fields(record: Dict[str, Any]) -> Dict[str, Any]:
    riepilogo = _extract_dettaglio_iva(record)
    ali = _aliquote_map(riepilogo)

    # Se nessuna aliquota è stata trovata (ali tutte None),
    # prova la scomposizione euristica per record atipici
    if all(v is None for v in ali.values()):
        imponibile = _extract_imponibile(record)
        imposta = parse_amount(record.get("impostaTotale"))
        if imponibile and imposta and imponibile > 0:
            scomposti = _scomponi_atipico(imponibile, imposta)
            for k, v in scomposti.items():
                if v is not None:
                    if k == "ESENTE":
                        ali["ESENTE"] = v
                    else:
                        ali[k] = v

    return {
        "id_invio": record.get("idInvio", ""),
        "matricola": record.get("matricolaDispositivo", ""),
        "tipo_dispositivo": record.get("tipoDispositivo", ""),
        "data_rilevazione": record.get("timeRilevazione", ""),
        "annullati": parse_amount(record.get("annullati")) or parse_amount(record.get("annullato")) or 0,
        "resi": parse_amount(record.get("resi")) or 0,
        "imponibile": _extract_imponibile(record),
        "imposta": parse_amount(record.get("impostaTotale")),
        "ali_4": ali["4"],
        "ali_5": ali["5"],
        "ali_10": ali["10"],
        "ali_22": ali["22"],
        "esente": ali["ESENTE"],
        "tipo_corrispettivo": "",
    }


def _identifica_problematici(
    all_records: List[Dict[str, Any]],
    raw_records: List[Dict[str, Any]],
    logger,
) -> set:
    """Individua i record con ripartizione IVA non-standard.

    Confronta il rapporto imposta/imponibile con le aliquote standard
    (4, 5, 10, 22) con tolleranza 0.1%. I record che non matchano
    ESATTAMENTE una standard hanno probabilmente aliquote miste e
    necessitano fallback browser.

    Returns:
        Set di idInvio (int) che necessitano fallback.
    """
    import json as _j
    problematici_ids: set = set()
    log_records: List[Dict] = []
    for raw, ext in zip(raw_records, all_records):
        imponibile = ext.get("imponibile", 0) or 0
        imposta = ext.get("imposta", 0) or 0
        id_invio = raw.get("idInvio")
        if id_invio and imponibile and imposta:
            ratio = imposta / imponibile * 100
            match_std = any(abs(ratio - std) < 0.1 for std in (4, 5, 10, 22))
            if not match_std:
                problematici_ids.add(int(id_invio) if str(id_invio).isdigit() else id_invio)
                log_records.append({
                    "idInvio": id_invio,
                    "ratio": round(ratio, 2),
                    "imponibile": raw.get("importoParzialeTotale"),
                    "imposta": raw.get("impostaTotale"),
                })
    if log_records:
        p = "/tmp/corrispettivi_problematici.json"
        try:
            with open(p, "w", encoding="utf-8") as f:
                _j.dump(log_records, f, ensure_ascii=False, indent=2, default=str)
        except Exception:
            pass
        logger(f"⚠️ {len(log_records)} record con rate atipica → fallback browser. File: {p}")
    return problematici_ids


def _fetch_dettaglio(
    session,
    headers_cons: Dict[str, str],
    dal_r: str,
    al_r: str,
    piva: str,
    tipo: str,
    unix_ms_func: Callable[[], str],
    logger,
) -> Optional[List[Dict[str, Any]]]:
    """Tenta di recuperare l'elenco corrispettivi con dettaglio IVA.

    Prova piu' pattern URL. Il pattern funzionante viene cachato
    (solo l'indice, non la stringa completa — cosi' le date vengono
    ricostruite correttamente ad ogni chunk).
    """
    base = "https://ivaservizi.agenziaentrate.gov.it/cons/cons-services/rs"

    # I pattern sono TEMPLATE con i placeholder {dal} e {al}.
    # La prima volta vengono costruiti, poi riusiamo l'indice del pattern vincente.
    if not hasattr(_fetch_dettaglio, "_working_idx"):
        _fetch_dettaglio._working_idx = None

    # Ricostruisci le URL con le date attuali
    date_params = f"dal/{dal_r}/al/{al_r}/piva/{piva}/tipoCorrispettivo/{tipo}"
    pattern_urls = [
        f"{base}/corrispettivi/dettaglio/{date_params}",
        f"{base}/corrispettivi/sintesi/elenco/{date_params}?dettaglio=true",
        f"{base}/corrispettivi/elenco/{date_params}",
        f"{base}/corrispettivi/riepilogo/{date_params}",
        f"{base}/corrispettivi/aliquote/{date_params}",
    ]

    if _fetch_dettaglio._working_idx is not None:
        patterns_to_try = [pattern_urls[_fetch_dettaglio._working_idx]]
    else:
        patterns_to_try = pattern_urls

    for idx, url in enumerate(patterns_to_try):
        url = f"{url}&v={unix_ms_func()}" if "?" in url else f"{url}?v={unix_ms_func()}"
        try:
            r = session.get(url, headers=headers_cons, timeout=15, verify=get_ca_bundle())
            if r.status_code == 200:
                data = r.json()
                rows = data.get("corrispettivi", []) if isinstance(data, dict) else data
                # DEBUG: salva risposta COMPLETA per ogni chunk
                import json as _j
                dp = f"/tmp/corrispettivi_full_response_{dal_r}_{al_r}.json"
                with open(dp, "w", encoding="utf-8") as _f:
                    _j.dump(data, _f, ensure_ascii=False, indent=2, default=str)
                logger(f"DEBUG: risposta completa salvata in {dp}")
                if rows and isinstance(rows, list) and len(rows) > 0:
                    if _fetch_dettaglio._working_idx is None:
                        _fetch_dettaglio._working_idx = idx
                        logger(f"Dettaglio: pattern {idx} funzionante! ({len(rows)} record)")
                    return rows
                # Lista vuota → fallback all'endpoint standard
                logger(f"Dettaglio pattern {idx}: HTTP 200 ma 0 record.")
            else:
                logger(f"Dettaglio pattern {idx}: HTTP {r.status_code}")
        except Exception as e:
            logger(f"Dettaglio pattern {idx}: errore — {e}")
            continue
    return None


def _fetch_dettaglio_record(
    session,
    headers_cons: Dict[str, str],
    id_invio: str,
    unix_ms_func: Callable[[], str],
    logger,
) -> Optional[Dict[str, Any]]:
    """Recupera il dettaglio righe per singolo corrispettivo (idInvio).

    Il portale mostra righe con Aliquota IVA, Imponibile, Imposta per
    ogni linea. L'endpoint dovrebbe restituire un array 'linee'/'righe'
    con la ripartizione per aliquota.
    """
    rs_base = "https://ivaservizi.agenziaentrate.gov.it/cons/cons-services/rs"
    brs_base = "https://ivaservizi.agenziaentrate.gov.it/cons/cons-services/brs"
    patterns = [
        # BRS (B2B REST - endpoint per dettaglio per-aliquota)
        f"{brs_base}/corrispettivi/{id_invio}",
        f"{brs_base}/corrispettivi/dettaglio/{id_invio}",
        f"{brs_base}/corrispettivi/sintesi/dettaglio/{id_invio}",
        # RS (standard REST)
        f"{rs_base}/corrispettivi/{id_invio}",
        f"{rs_base}/corrispettivi/sintesi/righe/{id_invio}",
        f"{rs_base}/corrispettivi/linee/{id_invio}",
    ]
    for i, url_template in enumerate(patterns):
        url = f"{url_template}?v={unix_ms_func()}"
        try:
            r = session.get(url, headers=headers_cons, timeout=15, verify=get_ca_bundle())
            if r.status_code == 200:
                data = r.json()
                logger(f"   Dettaglio record {id_invio}: pattern {i} OK")
                if not hasattr(_fetch_dettaglio_record, "_debug_saved"):
                    import json as _j
                    dp = "/tmp/corrispettivi_dettaglio_record_by_id.json"
                    with open(dp, "w", encoding="utf-8") as _f:
                        _j.dump(data, _f, ensure_ascii=False, indent=2, default=str)
                    logger(f"   DEBUG: dettaglio record salvato in {dp}")
                    _fetch_dettaglio_record._debug_saved = True
                return data if isinstance(data, dict) else None
        except Exception:
            continue

    return None


def fetch_elenco(
    session,
    headers_cons: Dict[str, str],
    dal: str,
    al: str,
    piva: str,
    unix_ms_func: Callable[[], str],
    logger,
    cf: str = "",
    pin: str = "",
    password: str = "",
) -> Tuple[List[Dict[str, Any]], Dict[int, Dict[str, Any]], List[Dict[str, Any]]]:
    """Scarica i corrispettivi dall'API e identifica quelli atipici.

    Returns:
        (records, id_atipici, raw_records_flat) dove:
        - records: lista record estratti
        - id_atipici: dict {idInvio: record_estratto} da aggiornare via browser
        - raw_records_flat: lista record raw (per debug)
    """
    base = "https://ivaservizi.agenziaentrate.gov.it/cons/cons-services/rs"
    brs_base = "https://ivaservizi.agenziaentrate.gov.it/cons/cons-services/brs"
    dal_r = date_to_rest(dal)
    al_r = date_to_rest(al)

    sintesi_url = (
        f"{base}/corrispettivi/sintesi/dal/{dal_r}/al/{al_r}/piva/{piva}?v={unix_ms_func()}"
    )
    r = session.get(sintesi_url, headers=headers_cons, verify=get_ca_bundle())
    if r.status_code != 200:
        logger(f"Corrispettivi: sintesi non disponibile (HTTP {r.status_code})")
        return []

    sintesi = r.json()
    totale_invii = int(sintesi.get("registratoriInvii", 0) or 0)
    logger(f"Corrispettivi: sintesi OK — registratoriInvii={totale_invii}")

    all_records: List[Dict[str, Any]] = []
    all_raw: List[Dict[str, Any]] = []
    id_atipici: Dict[int, Dict[str, Any]] = {}

    for tipo in TIPI_ELENCO:
        # Prova PRIMA l'endpoint BRS (B2B REST Service)
        # che restituisce i record PER-ALIQUOTA (non aggregati)
        brs_url = (
            f"{brs_base}/corrispettivi/elenco"
            f"/dal/{dal_r}/al/{al_r}/piva/{piva}/tipoCorrispettivo/{tipo}"
            f"?v={unix_ms_func()}"
        )
        r = session.get(brs_url, headers=headers_cons, verify=get_ca_bundle())
        if r.status_code == 200:
            data = r.json()
            rows = data.get("corrispettivi", data if isinstance(data, list) else [])
            if rows:
                logger(f"Corrispettivi [{tipo}]: {len(rows)} record (BRS per-aliquota)")
                msgs = []
                # Salva per debug
                import json as _j
                with open(f"/tmp/corrispettivi_brs_{tipo}.json", "w", encoding="utf-8") as _f:
                    _j.dump(data, _f, ensure_ascii=False, indent=2, default=str)
            else:
                logger(f"Corrispettivi [{tipo}]: BRS HTTP 200 ma 0 record")
                continue
        else:
            # Fallback: endpoint RS (standard - dati aggregati)
            url = (
                f"{base}/corrispettivi/sintesi/elenco"
                f"/dal/{dal_r}/al/{al_r}/piva/{piva}/tipoCorrispettivo/{tipo}"
                f"?v={unix_ms_func()}"
            )
            r = session.get(url, headers=headers_cons, verify=get_ca_bundle())
            if r.status_code != 200:
                logger(f"Corrispettivi [{tipo}]: HTTP {r.status_code} — saltato")
                continue
            data = r.json()
            rows = data.get("corrispettivi", [])
            msgs = data.get("messages", [])
        has_error = any(m.get("severity") == "ERROR" for m in msgs)

        if has_error:
            err_msg = "; ".join(m.get("message", "") for m in msgs if m.get("severity") == "ERROR")
            logger(f"Corrispettivi [{tipo}]: errore API — {err_msg}")
            continue

        if rows:
            logger(f"Corrispettivi [{tipo}]: {len(rows)} record trovati")
            raw_records = rows.copy()
            chunk_extracted: List[Dict[str, Any]] = []
            for rec in rows:
                ext = _extract_fields(rec)
                ext["tipo_corrispettivo"] = tipo
                chunk_extracted.append(ext)

            # Identifica atipici e raccoglili nel dict globale
            chunk_atipici = _identifica_problematici(chunk_extracted, raw_records, logger)
            for i, extracted in enumerate(chunk_extracted):
                idr = raw_records[i].get("idInvio") if i < len(raw_records) else None
                try:
                    idn = int(str(idr)) if idr is not None else None
                except (ValueError, TypeError):
                    idn = idr
                if idn and idn in chunk_atipici:
                    id_atipici[idn] = extracted

            all_records.extend(chunk_extracted)
            all_raw.extend(raw_records)

    logger(f"Corrispettivi: {len(all_records)} record totali, {len(id_atipici)} atipici")
    return all_records, id_atipici, all_raw


def export_csv(
    records: List[Dict[str, Any]],
    piva: str,
    dal: str,
    al: str,
    output_root: str,
    run_ts: str,
    anno: Optional[int] = None,
) -> str:
    # Nuova struttura Fase 2: output/{PIVA}/{ANNO}/corrispettivi/
    if anno is None:
        try:
            anno = int(dal.split("/")[2])
        except (ValueError, IndexError):
            anno = datetime.now().year
    folder = os.path.join(output_root, piva, str(anno), "corrispettivi")
    os.makedirs(folder, exist_ok=True)

    dal_fmt = dal.replace("/", "")
    al_fmt = al.replace("/", "")
    filename = f"corrispettivi_{dal_fmt}_{al_fmt}_{run_ts}.csv"
    path = os.path.join(folder, filename)

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "ID Invio", "Matricola dispositivo", "Tipo dispositivo",
            "Partita IVA", "Data rilevazione", "Annullati", "Resi",
            "Imponibile Giornata", "Imposta giornata",
            "Aliquota 4%", "Aliquota 5%", "Aliquota 10%", "Aliquota 22%", "Esente",
        ])
        for rec in records:
            writer.writerow([
                rec.get("id_invio", ""),
                rec.get("matricola", ""),
                rec.get("tipo_dispositivo", ""),
                piva,
                rec.get("data_rilevazione", ""),
                _fmt(rec.get("annullati")),
                _fmt(rec.get("resi")),
                _fmt(rec.get("imponibile")),
                _fmt(rec.get("imposta")),
                _fmt(rec.get("ali_4")),
                _fmt(rec.get("ali_5")),
                _fmt(rec.get("ali_10")),
                _fmt(rec.get("ali_22")),
                _fmt(rec.get("esente")),
            ])

    return path


def save_to_db(
    records: List[Dict[str, Any]],
    piva: str,
    logger,
) -> int:
    from app.database import get_session_factory, Corrispettivi, init_db

    init_db()
    Session = get_session_factory()
    db = Session()
    inserted = 0
    try:
        for rec in records:
            row = Corrispettivi(
                piva=piva,
                tipo_corrispettivo=rec.get("tipo_corrispettivo", ""),
                data_ora_rilevazione=rec.get("data_rilevazione"),
                imponibile_vendite=rec.get("imponibile"),
                imposta_vendite=rec.get("imposta"),
                id_invio=rec.get("id_invio", ""),
                matricola=rec.get("matricola", ""),
                importato_il=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            exists = (
                db.query(Corrispettivi)
                .filter_by(piva=piva, id_invio=rec.get("id_invio", ""))
                .first()
            )
            if not exists:
                db.add(row)
                inserted += 1
        db.commit()
    except Exception as e:
        db.rollback()
        logger(f"Corrispettivi DB: errore durante il salvataggio — {e}")
        raise
    finally:
        db.close()

    return inserted


def _safe_float(val: Any) -> Optional[float]:
    """Converte in float, restituisce None se non valido.

    Supporta sia già-float che stringhe formato IT (1.234,56) e EN (1234.56).
    Non restituisce mai stringhe o altri tipi — sempre float o None.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        return parse_amount(val)
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ─── Intestazioni fisse formato corrispettivi ──────────────────────────
# Colonne: ID Invio;Matricola dispositivo;Tipo dispositivo;Partita IVA;
#   Data rilevazione;Annullati;Resi;Imponibile Giornata;Imposta giornata;
#   Aliquota 4%;Aliquota 5%;Aliquota 10%;Aliquota 22%;Esente;
_RECU_HEADERS = [
    "ID Invio",
    "Matricola dispositivo",
    "Tipo dispositivo",
    "Partita IVA",
    "Data rilevazione",
    "Annullati",
    "Resi",
    "Imponibile Giornata",
    "Imposta giornata",
    "Aliquota 4%",
    "Aliquota 5%",
    "Aliquota 10%",
    "Aliquota 22%",
    "Esente",
]


def export_xlsx(
    records: List[Dict[str, Any]],
    piva: str,
    dal: str,
    al: str,
    output_root: str,
    run_ts: str,
    anno: Optional[int] = None,
) -> str:
    """
    Esporta i corrispettivi in formato Excel (.xlsx).

    Colonne fisse: ID Invio, Matricola dispositivo, Tipo dispositivo,
    Partita IVA, Data rilevazione, Annullati, Resi,
    Imponibile Giornata, Imposta giornata,
    Aliquota 4%, Aliquota 5%, Aliquota 10%, Aliquota 22%, Esente.

    I valori nelle colonne aliquote sono presi dal 'riepilogo' di ogni
    record e incasellati nella colonna corrispondente all'aliquota IVA
    applicata. Numeri in formato #,##0.00.
    """
    # Assicura che _libs sia nel path per openpyxl
    import sys
    _libs_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "_libs")
    if os.path.isdir(_libs_path) and _libs_path not in sys.path:
        sys.path.insert(0, _libs_path)

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if anno is None:
        try:
            anno = int(dal.split("/")[2])
        except (ValueError, IndexError):
            anno = datetime.now().year

    folder = os.path.join(output_root, piva, str(anno), "corrispettivi")
    os.makedirs(folder, exist_ok=True)

    dal_fmt = dal.replace("/", "")
    al_fmt = al.replace("/", "")
    filename = f"corrispettivi_{dal_fmt}_{al_fmt}_{run_ts}.xlsx"
    path = os.path.join(folder, filename)

    headers = _RECU_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = "Corrispettivi"

    # Intestazioni in grassetto
    bold_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = bold_font

    # Colonne con formato #,##0.00: Annullati=6, Resi=7, Imponibile=8,
    # Imposta=9, Aliquota 4%=10, 5%=11, 10%=12, 22%=13, Esente=14
    _DECIMAL_COLS = {6, 7, 8, 9, 10, 11, 12, 13, 14}

    for row_idx, rec in enumerate(records, 2):
        # ID Invio come numero
        try:
            id_invio = int(rec.get("id_invio", 0))
        except (ValueError, TypeError):
            id_invio = rec.get("id_invio", "")

        # Partita IVA come numero
        try:
            piva_num = int(piva)
        except (ValueError, TypeError):
            piva_num = piva

        # Data rilevazione come datetime
        data_raw = rec.get("data_rilevazione", "")
        if data_raw and isinstance(data_raw, str):
            try:
                data_rilevazione = datetime.fromisoformat(data_raw.replace("Z", "+00:00"))
            except (ValueError, TypeError):
                data_rilevazione = data_raw
        else:
            data_rilevazione = data_raw

        # I valori aliquote sono già estratti da _extract_fields come
        # ali_4, ali_5, ali_10, ali_22, esente (imponibile + imposta)
        values: List[Any] = [
            id_invio,
            rec.get("matricola", ""),
            rec.get("tipo_dispositivo", ""),
            piva_num,
            data_rilevazione,
            _safe_float(rec.get("annullati")),
            _safe_float(rec.get("resi")),
            _safe_float(rec.get("imponibile")),
            _safe_float(rec.get("imposta")),
            _safe_float(rec.get("ali_4")),
            _safe_float(rec.get("ali_5")),
            _safe_float(rec.get("ali_10")),
            _safe_float(rec.get("ali_22")),
            _safe_float(rec.get("esente")),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in _DECIMAL_COLS:
                cell.number_format = '#,##0.00'
            elif col in (1, 4) and isinstance(val, int):
                # ID Invio e Partita IVA come numeri interi (no decimali)
                cell.number_format = '0'

    # Larghezza colonne
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 2, 14)

    wb.save(path)
    return path


def run(
    session,
    headers_cons: Dict[str, str],
    dal: str,
    al: str,
    piva: str,
    cfg: Dict[str, str],
    unix_ms_func: Callable[[], str],
    logger,
    output_root: str = "output",
    run_ts: Optional[str] = None,
    anno: Optional[int] = None,
    cf: str = "",
    pin: str = "",
    password: str = "",
) -> bool:
    if run_ts is None:
        run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger(f"\nCorretta avvio download corrispettivi [{dal} → {al}]...")

    chunks = get_date_chunks(dal, al)
    all_records: List[Dict[str, Any]] = []
    all_atipici: Dict[int, Dict[str, Any]] = {}

    for d_dal, d_al in chunks:
        try:
            records, atipici, _ = fetch_elenco(
                session, headers_cons, d_dal, d_al, piva, unix_ms_func, logger,
                cf=cf, pin=pin, password=password,
            )
            all_records.extend(records)
            all_atipici.update(atipici)
        except Exception as e:
            logger(f"Corrispettivi chunk [{d_dal} → {d_al}]: errore — {e}")

    # Nota: gli atipici sono già gestiti da _extract_fields con _scomponi_atipico.
    # Il browser fallback è stato rimosso in favore dell'euristico.
    if all_atipici:
        logger(f"Note: {len(all_atipici)} record atipici risolti via euristico")

    if not all_records:
        logger("Corrispettivi: nessun record da salvare.")
        return True

    try:
        xlsx_path = export_xlsx(all_records, piva, dal, al, output_root, run_ts, anno=anno)
        logger(f"Corrispettivi: Excel salvato in {xlsx_path} ({len(all_records)} righe)")
    except Exception as e:
        logger(f"Corrispettivi: errore export Excel — {e}")
        # Fallback a CSV se Excel fallisce
        try:
            csv_path = export_csv(all_records, piva, dal, al, output_root, run_ts, anno=anno)
            logger(f"Corrispettivi: fallback CSV salvato in {csv_path}")
        except Exception as e2:
            logger(f"Corrispettivi: anche fallback CSV fallito — {e2}")
            return False

    if cfg.get("WRITE", "1") == "0":
        try:
            if xlsx_path:
                os.remove(xlsx_path)
            logger("Corrispettivi WRITE=0: rimosso file Excel")
        except Exception:
            pass

    if cfg.get("DB", "0") == "1":
        try:
            n = save_to_db(all_records, piva, logger)
            logger(f"Corrispettivi DB: {n} nuovi record inseriti su {len(all_records)} totali")
        except Exception as e:
            logger(f"Corrispettivi DB: errore — {e}")

    return True
